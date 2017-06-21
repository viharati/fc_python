from openpyxl import Workbook
#class variable wb
wb=Workbook()

ws=wb.create_sheet('sheet_test')
ws['A1']='a1-1'
ws['B1']='b1-1'
ws['C1']='=1+1'
ws['d1']='=1+1'

wb.save('my_result.xlsx')
