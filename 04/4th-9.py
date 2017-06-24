from openpyxl import load_workbook
wb = load_workbook('my_result.xlsx', read_only=True)
data = wb['Sheet']

for row in data.iter_rows(min_row=2, min_col=2, max_col=5, max_row=7):
#for row in data.iter_rows():
#for row in data.iter_cols(): # no attribute error

	for cell in row:
		print(cell.value)