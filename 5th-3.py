from openpyxl import load_workbook
from my_email import Email
from my_email_multi import Email as EmailMulti

wb = load_workbook('emails.xlsx')
ws = wb['Sheet1']

e=Email()
#e=EmailMulti()

for row in ws.iter_rows():
	e.send_mail(row[0].value.encode('utf-8'), row[1].value, 'emails.xlsx')
	print(row[0].value, "\n" ,row[1].value,)


print('-'*10)

row_cnt =1
while True:
	if not ws['A'+str(row_cnt)].value:
		break

	name = ws['A'+str(row_cnt)].value
	addr = ws['B'+str(row_cnt)].value
	e.send_mail(name.encode('utf-8'), addr)

	print(name, "\n" ,addr)
	row_cnt = row_cnt + 1