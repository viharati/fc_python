from openpyxl import Workbook
from openpyxl import load_workbook
#wb=Workbook()
#wb=wb.load_workbook('sample.xlsx')



wb=load_workbook('sample.xlsx')
ws=wb['Sheet1']
wr=wb['Sheet2']
##
write_wb=Workbook()
write_wb.create_sheet('Sheet0')
write_sheet0=write_wb['Sheet0']

write_wb.create_sheet('Sheet1')
write_sheet1=write_wb['Sheet1']


idx = 0 
while True:
	idx = idx + 1
	row = ws[idx]
	if 'study' in str(row[3].value):
		print (row[3].value)
		wr.append([row[3].value])
		##
		write_sheet0.append([row[3].value])
		write_sheet1.append([row[].value])


		break

wb.save('result.xlsx')
write_wb.save('result2.xlsx')
